from time import localtime, strftime
from itertools import takewhile, repeat
from glob import glob
from tqdm import tqdm
import openpyxl
import os

XL_OUTPUT_DIR_HOME = 'QuartXL-Report'

forbidden_char_filename = {
    # src https://stackoverflow.com/questions/1976007/what-characters-are-forbidden-in-windows-and-linux-directory-names
    '/' : chr(0x2215),
    '\\' : chr(0x29F5),
    ':' : chr(0xA789),
    '*' : chr(0x2217),
    '?' : chr(0xFE56),
    '"' : chr(0x201C),
    '<' : chr(0xFE64),
    '>' : chr(0xFE65),
    '|' : chr(0x01C0)
}

class Parser:
    IDLE = 0
    INITIALIZE = 1
    COLUMN = 2
    READING = 3
    FINISHED = 4
    
    def __init__(self):
        self.reset()
        self.__map = {
            Parser.INITIALIZE : self.__set_table_name,
            Parser.COLUMN : self.__set_columns,
            Parser.READING : self.__process_data
        }
    
    def reset(self):
        self.__current_state = Parser.IDLE
        self.__column_called = False #Some tables don't have proper column
    
    def process_line(self, line:str):
        if(line.startswith('+-')):
            self.__current_state += 1
            return
        if(not line.startswith(';')):
            self.reset()
            return
        func = self.__map.get(self.__current_state)
        if(func):
            func(line)
        
    
    def __set_table_name(self,line:str):
        self.__table_name = line.strip().strip(';').strip()
    
    def __set_columns(self, line:str):
        if(self.__column_called):
            self.__no_proper_column(line)
            return
        __columns = line.strip().strip(';').split(';')
        _columns = dict()
        length = []
        for i in __columns:
            c_name = i.strip()
            _columns[c_name]= []
            length.append(len(c_name))
        
        self.__columns = _columns
        self.__length = length
        self.__column_called = True
    
    def __no_proper_column(self,line):
        self.__current_state = Parser.READING
        __columns = dict()
        for i, col in enumerate(self.__columns, start=1):
            __columns['_' * i] = [col]
        self.__columns = __columns
        self.__process_data(line)
    
    def __process_data(self, line:str):
        data = line.strip().strip(';').split(';')
        i = 0
        for c,d in zip(self.__columns,data):
            value = d.strip()
            self.__columns[c].append(value)
            self.__length[i] = max(self.__length[i],len(value))
            i += 1

    
    def get_current_state(self):
        return self.__current_state
    
    def get_json(self):
        ret = dict()
        ret['table_name'] = self.__table_name
        ret['data'] = self.__columns
        ret['length'] = self.__length
        
        self.reset()
        return ret

def replace_forbidden(name:str):
    for i in forbidden_char_filename:
        name = name.replace(i,forbidden_char_filename[i])
    return name

def save_xl(json, save_path):
    file_name = replace_forbidden(json['table_name']) + '.xlsx'
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    for i, column in enumerate(json['data'], start=1):
        ws_column = sheet.cell(row = 1, column = i).column_letter
        sheet.column_dimensions[ws_column].width = (json['length'][i-1] + 4) * 1.2
        if(column.replace('_','')):
            cell = sheet.cell(row = 1, column = i)
            cell.font = openpyxl.styles.Font(name='Courier New',bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            cell.value = column
            col_start = 2
        else:
            col_start = 1
        
        for j, value in enumerate(json['data'][column],start=col_start):
            cell = sheet.cell(row = j, column = i)
            cell.font = openpyxl.styles.Font(name='Courier New')
            cell.value = value
    
    file_path = os.path.join(save_path,file_name)
    i = 1
    while(os.path.exists(file_path)):
        file_path = os.path.join(save_path,file_name.split('.')[-2] + '-' + str(i)+'.xlsx')
        i += 1
    
    workbook.save(file_path)
    workbook.close()

def get_line_count(path):
    # Copied from: https://stackoverflow.com/questions/845058/how-to-get-line-count-of-a-large-file-cheaply-in-python
    # Answer by: Michael Bacon ( https://stackoverflow.com/users/4367773/michael-bacon )

    with open(path, 'rb') as f:
        bufgen = takewhile(lambda x: x, (f.raw.read(1024*1024) for _ in repeat(None)))
        return sum(buf.count(b'\n') for buf in bufgen)

def main():
    reports = [i for i in glob('output_files/*.rpt')]
    if(not reports):
        print('No reports found. Make sure you have compiled using Quartus Prime and you are on correct diretory.')
        return
    
    run_dir = strftime('%Y-%m-%d %I:%M:%S%p', localtime()).lower()
    run_dir = replace_forbidden(run_dir)
    run_dir = os.path.join(XL_OUTPUT_DIR_HOME, run_dir)
    os.makedirs(run_dir, exist_ok=True)

    parser = Parser()

    for r in reports:
        parser.reset()
        total_line = get_line_count(r)

        file = open(r, 'r', encoding='utf-8',errors='backslashreplace')
        try:
            report_name = file.readline().split(' report ')[0]
        except:
            report_name = os.path.basename(r)
        print('Genarating Excel Report for', report_name)

        save_path = os.path.join(run_dir, report_name)
        if(not os.path.exists(save_path)):
            os.mkdir(save_path)
        
        for line in tqdm(file,total=total_line-1, unit='line'):
            parser.process_line(line)
            if(parser.get_current_state()==Parser.FINISHED):
                json = parser.get_json()
                save_xl(json, save_path)
                
        file.close()


if __name__ == '__main__':
    main()
